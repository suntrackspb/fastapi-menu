import hashlib
import json
from datetime import timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from celery import Celery
from openpyxl import load_workbook
from sqlalchemy import create_engine

from app.config import DB_HOST, DB_NAME, DB_PASS, DB_PORT, DB_USER, USE_GOOGLE
from app.google_sheets.google_sheet import auth, get_data

engine = create_engine(f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

celery_app = Celery()

celery_app.conf.broker_url = "amqp://guest:guest@rabbitmq:5672//"
celery_app.conf.result_backend = "rpc://"
celery_app.conf.broker_connection_retry_on_startup = True

celery_app.conf.beat_schedule = {
    "pandas_update_database": {
        "task": "app.celery.tasks.pandas_update_database",
        "schedule": timedelta(seconds=15),
    },
}

admin_file = Path("./app/admin/Menu.xlsx")
hash_file = Path("./app/admin/hash")


def calculate_file_hash() -> str:
    with Path("./app/admin/Menu.xlsx").open("rb") as f:
        hasher = hashlib.sha256()
        while chunk := f.read(65536):
            hasher.update(chunk)
    return hasher.hexdigest()


def read_hash() -> str:
    with Path("./app/admin/hash").open("r") as f:
        return f.read()


def write_hash(hash_summ: str) -> None:
    with Path("./app/admin/hash").open("w") as f:
        f.write(hash_summ)


def excel_to_json(array: list[list]) -> tuple[dict[str, dict[Any, Any]], dict[str, dict[Any, Any]], dict[str, dict[Any, Any]]]:

    menu_json: dict[str, dict] = {"id": {}, "title": {}, "description": {}}
    submenu_json: dict[str, dict] = {"id": {}, "menu_id": {}, "title": {}, "description": {}}
    dish_json: dict[str, dict] = {"id": {}, "submenu_id": {},
                                  "title": {}, "description": {}, "price": {}, "discount": {}}

    menu_count, sub_count, dish_count = 0, 0, 0

    current_menu_id = ""
    current_sub_id = ""

    for row in array:
        if bool(row[0]) and bool(row[1]):
            current_menu_id = row[0]
            menu_json["id"][menu_count] = row[0]
            menu_json["title"][menu_count] = row[1]
            menu_json["description"][menu_count] = row[2]
            menu_count += 1

        elif bool(row[0]) is False and bool(row[1]):
            current_sub_id = row[1]
            submenu_json["id"][sub_count] = row[1]
            submenu_json["menu_id"][sub_count] = current_menu_id
            submenu_json["title"][sub_count] = row[2]
            submenu_json["description"][sub_count] = row[3]
            sub_count += 1

        elif bool(row[0]) is False and bool(row[1]) is False:
            dish_json["id"][dish_count] = row[2]
            dish_json["submenu_id"][dish_count] = current_sub_id
            dish_json["title"][dish_count] = row[3]
            dish_json["description"][dish_count] = row[4]
            dish_json["price"][dish_count] = row[5]
            dish_json["discount"][dish_count] = 0
            if len(row) > 6 and row[6] is not None:
                dish_json["discount"][dish_count] = row[6]
            dish_count += 1
    return menu_json, submenu_json, dish_json


def run_update_database(data: list) -> None:
    print("===" * 30)
    print("RUN UPDATE DATABASE")
    print("===" * 30)

    menus_data, submenus_data, dishes_data = excel_to_json(data)
    dish_df = pd.read_json(json.dumps(dishes_data))
    dish_df.to_sql("dishes", engine, if_exists="replace", index=False)

    submenu_df = pd.read_json(json.dumps(submenus_data))
    submenu_df.to_sql("submenus", engine, if_exists="replace", index=False)

    menu_df = pd.read_json(json.dumps(menus_data))
    menu_df.to_sql("menus", engine, if_exists="replace", index=False)


@celery_app.task
def pandas_update_database() -> None:
    bool_value = (USE_GOOGLE == "True")
    if bool_value:
        creds = auth()
        data = get_data(creds)
        print("===" * 15)
        print("USE GOOGLE")
        print("===" * 15)
        run_update_database(data)
    else:
        if Path(admin_file).exists():
            new_hash = calculate_file_hash()
            if not Path.exists(hash_file):
                write_hash("22222")
            old_hash = read_hash()
            wb = load_workbook(admin_file)
            sheet = wb.active
            data = sheet.iter_rows(values_only=True)
            if old_hash != new_hash:
                print("===" * 15)
                print("USE LOCAL")
                print("===" * 15)
                run_update_database(data)
                write_hash(new_hash)
            else:
                print("===" * 15)
                print("DIFF NOT FOUND")
                print("===" * 15)
