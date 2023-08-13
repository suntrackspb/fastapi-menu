from pathlib import Path
from typing import Any

import httpx
import openpyxl
import openpyxl.styles

from app.config import BASE_URL, EXCEL_FILE
from app.crud.data import DataCrud


class DataService:
    def __init__(self, crud: DataCrud):
        self.crud = crud

    async def get_full_with_id(self) -> list[dict]:
        return await self.crud.get_list_with_id()

    async def get_full_without_id(self) -> list[dict[Any, Any]] | None:
        return await self.crud.get_list_without_id()

    # DEBUG FUNCTIONS

    @staticmethod
    async def convert_xls_to_json(filepath: Path) -> list[dict]:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        json_obj = []
        menu: dict = {}
        submenu: dict = {}

        for row in sheet.iter_rows(values_only=True):
            if row[0] is not None and row[1] is not None:
                menu = {
                    "title": row[1],
                    "description": row[2],
                    "submenus": [],
                }
                json_obj.append(menu)
            if row[0] is None and row[1] is not None:
                submenu = {
                    "title": row[2],
                    "description": row[3],
                    "dishes": [],
                }
                menu["submenus"].append(submenu)
            if row[0] is None and row[1] is None:
                dish = {
                    "title": row[3],
                    "description": row[4],
                    "price": str(row[5]),
                }
                submenu["dishes"].append(dish)
        return json_obj

    @staticmethod
    async def form_excel(extracted_data: Any) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Menus"
        style = openpyxl.styles.NamedStyle(name="bold")
        style.font = openpyxl.styles.Font(bold=True)

        row = 0
        for data in extracted_data:
            sheet.cell(row=row + 1, column=1, value=str(data.id))
            sheet.cell(row=row + 1, column=2, value=data.title).style = style
            sheet.cell(row=row + 1, column=3, value=data.description).style = style
            row += 1
            for submenu in data.submenus:
                sheet.cell(row=row + 1, column=2, value=str(submenu.id))
                sheet.cell(row=row + 1, column=3, value=submenu.title).style = style
                sheet.cell(row=row + 1, column=4, value=submenu.description).style = style
                row += 1
                for dish in submenu.dishes:
                    sheet.cell(row=row + 1, column=3, value=str(dish.id))
                    sheet.cell(row=row + 1, column=4, value=dish.title).style = style
                    sheet.cell(row=row + 1, column=5, value=dish.description).style = style
                    sheet.cell(row=row + 1, column=6, value=dish.price).style = style
                    row += 1
        workbook.save("./admin/Database.xlsx")

    @staticmethod
    async def upload_to_database(obj: list[dict]) -> None:
        async with httpx.AsyncClient() as client:
            for menu in obj:
                response_menu = await client.post(f"{BASE_URL}/menus", json=menu)
                db_menu = response_menu.json()
                for submenu in menu["submenus"]:
                    response_submenu = await client.post(
                        f"{BASE_URL}/menus/{db_menu['id']}/submenus",
                        json=submenu,
                    )
                    db_sub = response_submenu.json()
                    for dish in submenu["dishes"]:
                        await client.post(
                            f"{BASE_URL}/menus/{db_menu['id']}/submenus/{db_sub['id']}/dishes",
                            json=dish,
                        )

    async def load_to_database(self) -> dict[str, str]:
        json_data = await self.convert_xls_to_json(EXCEL_FILE)
        await self.upload_to_database(json_data)
        return {"status": "true", "message": "Import successful"}

    async def unload_to_excel(self) -> dict[str, str]:
        db_data = await self.crud.get_list_with_id()
        await self.form_excel(db_data)
        return {"status": "true", "message": "Import successful"}
