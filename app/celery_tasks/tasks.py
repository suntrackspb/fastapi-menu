import hashlib
import json
from datetime import timedelta
from pathlib import Path
from typing import Any

import gspread
import pandas as pd
from celery import Celery
from openpyxl import load_workbook
from sqlalchemy import Integer, String, create_engine
from sqlalchemy.dialects.postgresql import UUID

from app.config import (
    DB_HOST,
    DB_NAME,
    DB_PASS,
    DB_PORT,
    DB_USER,
    SHEET_UID,
    USE_GOOGLE,
)

engine = create_engine(f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

celery_app = Celery()

celery_app.conf.broker_url = "amqp://guest:guest@rabbitmq:5672//"
celery_app.conf.result_backend = "rpc://"
celery_app.conf.broker_connection_retry_on_startup = True

celery_app.conf.beat_schedule = {
    "pandas_update_database": {
        "task": "app.celery_tasks.tasks.pandas_update_database",
        "schedule": timedelta(seconds=15),
    },
}

admin_file = Path("./app/admin/Menu.xlsx")
hash_file = Path("./app/admin/hash")


def calculate_file_hash() -> str:
    with Path("./app/admin/Menu.xlsx").open("rb") as f:
        hasher = hashlib.sha256()
        while chunk := f.read(65536):
            hasher.update(chunk)
    return hasher.hexdigest()


def read_hash() -> str:
    with Path("./app/admin/hash").open("r") as f:
        return f.read()


def write_hash(hash_summ: str) -> None:
    with Path("./app/admin/hash").open("w") as f:
        f.write(hash_summ)


def excel_to_json(array: list[list]) -> tuple[dict[str, dict[Any, Any]], dict[str, dict[Any, Any]], dict[str, dict[Any, Any]]]:
    menu_json: dict[str, dict] = {"id": {}, "title": {}, "description": {}}
    submenu_json: dict[str, dict] = {"id": {}, "menu_id": {}, "title": {}, "description": {}}
    dish_json: dict[str, dict] = {"id": {}, "submenu_id": {},
                                  "title": {}, "description": {}, "price": {}, "discount": {}}

    menu_count, sub_count, dish_count = 0, 0, 0

    current_menu_id = ""
    current_sub_id = ""

    for row in array:
        if bool(row[0]) and bool(row[1]):
            current_menu_id = row[0]
            menu_json["id"][menu_count] = row[0]
            menu_json["title"][menu_count] = row[1]
            menu_json["description"][menu_count] = row[2]
            menu_count += 1

        elif bool(row[0]) is False and bool(row[1]):
            current_sub_id = row[1]
            submenu_json["id"][sub_count] = row[1]
            submenu_json["menu_id"][sub_count] = current_menu_id
            submenu_json["title"][sub_count] = row[2]
            submenu_json["description"][sub_count] = row[3]
            sub_count += 1

        elif bool(row[0]) is False and bool(row[1]) is False:
            dish_json["id"][dish_count] = row[2]
            dish_json["submenu_id"][dish_count] = current_sub_id
            dish_json["title"][dish_count] = row[3]
            dish_json["description"][dish_count] = row[4]
            dish_json["price"][dish_count] = row[5]
            dish_json["discount"][dish_count] = 0
            if len(row) > 6 and bool(row[6]):
                dish_json["discount"][dish_count] = row[6]
            dish_count += 1
    return menu_json, submenu_json, dish_json


def run_update_database(data: list) -> None:
    menus_data, submenus_data, dishes_data = excel_to_json(data)
    dish_df = pd.read_json(json.dumps(dishes_data))
    dish_df.to_sql("dishes", engine, if_exists="replace", index=False,
                   dtype={"id": UUID(as_uuid=True),
                          "submenu_id": UUID(as_uuid=True),
                          "title": String,
                          "description": String,
                          "price": String,
                          "discount": Integer,
                          },
                   )

    submenu_df = pd.read_json(json.dumps(submenus_data))
    submenu_df.to_sql("submenus", engine, if_exists="replace", index=False,
                      dtype={"id": UUID(as_uuid=True),
                             "menu_id": UUID(as_uuid=True),
                             "title": String,
                             "description": String,
                             },
                      )

    menu_df = pd.read_json(json.dumps(menus_data))
    menu_df.to_sql("menus", engine, if_exists="replace", index=False,
                   dtype={"id": UUID(as_uuid=True),
                          "title": String,
                          "description": String,
                          },
                   )


@celery_app.task
def pandas_update_database() -> None:
    bool_value = (USE_GOOGLE == "True")
    if bool_value:
        gc = gspread.service_account(filename="./credentials.json")
        last_update = gc.open_by_key(SHEET_UID).lastUpdateTime
        if not Path.exists(hash_file):
            write_hash("22222")
        old_time = read_hash()
        if last_update != old_time:
            sh = gc.open_by_key(SHEET_UID)
            worksheet = sh.get_worksheet(0)
            list_of_lists = worksheet.get_all_values()
            run_update_database(list_of_lists)
            write_hash(last_update)
    else:
        if Path(admin_file).exists():
            new_hash = calculate_file_hash()
            if not Path.exists(hash_file):
                write_hash("22222")
            old_hash = read_hash()
            wb = load_workbook(admin_file)
            sheet = wb.active
            data = sheet.iter_rows(values_only=True)
            if old_hash != new_hash:
                run_update_database(data)
                write_hash(new_hash)
